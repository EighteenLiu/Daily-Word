from __future__ import annotations

import re
import random
import struct
import zipfile
import bisect
import hashlib
from io import BytesIO
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Iterable
import xml.etree.ElementTree as ET


COMMUNITY_CATEGORY = "居住小区、平房胡同"
RESTAURANT_CATEGORY = "餐饮单位"
SOCIAL_UNIT_CATEGORY = "社会单位"
FALLBACK_IMAGE_COLUMNS = set(range(8, 26))
OUTSIDE_BUCKET_POINT = "厨余、其他垃圾桶外摆检查"
OUTSIDE_BUCKET_PROBLEM_PHOTO_COLUMNS = set(range(8, 19))

UNIT_INDICATOR_CHILDREN: dict[str, dict[str, tuple[str, ...]]] = {
    RESTAURANT_CATEGORY: {
        "良好，未发现问题": (),
        "宣传引导情况": ("无垃圾分类投放指引", "无四周年宣传海报", "无宣传氛围", "无源头减量宣传（光盘行动）"),
        "容器品类成组设置": ("后厨未成组设置垃圾桶", "无容器设置", "成组配置", "集中用餐区未成组设置容器"),
        "容器标识": ("良好，未发现问题", "容器标识不正确或无标识", "容器标识不清晰、破损", "容器颜色不正确"),
        "容器检查": ("容器垃圾不纯净", "良好，未发现问题", "投放点外表面明显脏污", "桶站破损、脏污", "桶站周边不洁", "桶站满冒", "站外摆桶"),
        "垃圾收运合同": ("无厨余垃圾收运合同或不合格", "无其他垃圾收运合同或不合格"),
        "垃圾排放登记": ("无厨余垃圾排放登记方式", "无非居民其他垃圾排放登记方式"),
        "垃圾桶“游街”": ("有", "无"),
        "厨余垃圾桶外摆": ("有", "无"),
        "无油水分离装置": ("有", "无"),
        "无废弃油脂合同或不合格": (),
        "无隔油池": (),
        "可回收物和其他垃圾乱堆乱放": ("有", "无"),
        "主动提供一次性用品": ("是", "否"),
        "分类投放情况": ("投放正确", "投放错误"),
        "无防蝇灭鼠设备": ("有",),
        "无称重计量小程序": (),
        "无收费计量记录": (),
        "无源头减量措施": (),
        "无便利性措施": (),
        "未更新新国际": (),
    },
    SOCIAL_UNIT_CATEGORY: {
        "良好，未发现问题": (),
        "宣传引导情况": ("有", "无", "无宣传氛围", "无源头减量宣传（光盘行动）", "无四周年宣传海报"),
        "无党建引领相关资料": (),
        "无培训活动、会议记录、照片等培训材料": (),
        "无四分类垃圾清运台账或四分类清运台账不合格": (),
        "容器品类设置": ("单位无容器配置", "公共场所区域未成组设置可回收物和其他容器", "集中用餐区未成组设置厨余和其他容器", "食品加工区未成组设置厨余和其他容器", "成组配置", "未成组配置"),
        "容器标识": ("良好，未发现问题", "容器标识不正确", "容器标识不清晰、破损", "容器颜色不正确"),
        "投放点环境（容器检查）": ("良好，未发现问题", "投放点外表面明显脏污", "桶站破损、脏污", "桶站周边不洁", "桶站满冒", "站外摆桶"),
        "垃圾桶“游街”": ("有", "无"),
        "厨余垃圾桶外摆": ("有", "无"),
        "可回收物和其他垃圾乱堆乱放": ("有", "无"),
        "分类投放情况": ("投放正确", "投放错误"),
        "未设置分类投放指引数（组数）": (),
        "容器无便利性措施": (),
        "收运合同": ("无厨余垃圾收运合同或不合格", "无其他垃圾收运合同或不合格", "无可回收物收运合同或不合格", "无有害垃圾收运合同或不合格"),
        "排放登记": ("无厨余垃圾排放登记方式", "无非居民其他垃圾排放登记方式"),
        "无源头减量措施": (),
        "无废弃油脂合同或不合格": (),
        "无垃圾分类工作方案": (),
        "职责分工不明确": (),
        "无称重计量列表": (),
        "无油水分离装置": (),
    },
}

UNIT_INDICATOR_ALIASES: dict[str, dict[str, str]] = {
    RESTAURANT_CATEGORY: {
        "宣传引导": "宣传引导情况",
        "容器成组设置": "容器品类成组设置",
        "油水分离装置": "无油水分离装置",
    },
    SOCIAL_UNIT_CATEGORY: {
        "宣传引导": "宣传引导情况",
        "容器成组设置": "容器品类设置",
        "投放点环境": "投放点环境（容器检查）",
        "油水分离装置": "无油水分离装置",
    },
}

UNIT_NO_PROBLEM_RESULTS: dict[str, dict[str, set[str]]] = {
    RESTAURANT_CATEGORY: {
        "宣传引导情况": {"有", "有宣传氛围"},
        "容器品类成组设置": {"成组配置"},
        "容器标识": {"良好，未发现问题", "良好,未发现问题"},
        "容器检查": {"良好，未发现问题", "良好,未发现问题"},
        "垃圾桶“游街”": {"无"},
        "厨余垃圾桶外摆": {"无"},
        "无油水分离装置": {"有"},
        "无防蝇灭鼠设备": {"有"},
        "可回收物和其他垃圾乱堆乱放": {"无"},
        "主动提供一次性用品": {"否"},
        "分类投放情况": {"投放正确"},
    },
    SOCIAL_UNIT_CATEGORY: {
        "宣传引导情况": {"有", "有宣传氛围"},
        "容器品类设置": {"成组配置"},
        "容器标识": {"良好，未发现问题", "良好,未发现问题"},
        "投放点环境（容器检查）": {"良好，未发现问题", "良好,未发现问题"},
        "垃圾桶“游街”": {"无"},
        "厨余垃圾桶外摆": {"无"},
        "可回收物和其他垃圾乱堆乱放": {"无"},
        "分类投放情况": {"投放正确"},
        "无油水分离装置": {"有"},
    },
}

UNIT_NO_PROBLEM_TEXT_MARKERS: dict[str, dict[str, set[str]]] = {
    RESTAURANT_CATEGORY: {
        "宣传引导情况": {"有宣传氛围"},
        "无油水分离装置": {"有", "有油水分离装置"},
        "无防蝇灭鼠设备": {"有", "有防蝇灭鼠设备", "有防蝇设备", "有灭鼠设备"},
        "主动提供一次性用品": {"否", "未主动提供一次性用品", "未提供一次性用品"},
    },
    SOCIAL_UNIT_CATEGORY: {
        "宣传引导情况": {"有宣传氛围"},
        "无油水分离装置": {"有", "有油水分离装置"},
    },
}

UNIT_IGNORED_INDICATORS: dict[str, set[str]] = {
    RESTAURANT_CATEGORY: {"主动提供一次性用品"},
}

UNIT_GENERIC_NO_PROBLEM_RESULTS = {"", "无问题", "良好，未发现问题", "良好,未发现问题", "未发现问题"}
UNIT_GENERIC_PROBLEM_RESULTS_USE_PARENT = {"有", "是", "无"}
UNIT_POSITIVE_RESULT_VALUES = {"有", "是", "已设置", "已配置", "已更新", "正常", "合格"}
UNIT_NEGATIVE_RESULT_VALUES = {"无", "否", "未发现", "未提供"}
UNIT_BAD_EVENT_INDICATOR_KEYWORDS = ("游街", "外摆", "乱堆乱放")
PROBLEM_KEYWORDS = (
    "不洁",
    "不准确",
    "站外摆桶",
    "桶外摆",
    "外摆",
    "满冒",
    "破损",
    "脏污",
    "不齐全",
    "无宣传",
    "未开盖",
    "混投",
    "散桶",
)

COMMUNITY_EXISTENCE_INDICATOR_PROBLEM_TEXT = {
    "大件垃圾投放点": "无大件垃圾投放点",
    "装修垃圾投放点": "无装修垃圾投放点",
}
COMMUNITY_SETTING_INDICATOR_SUFFIXES = {
    "大件垃圾投放点设置": (
        "公示牌信息错误",
        "无大件垃圾托底上门回收信息",
        "未设置公示牌",
        "未设置围挡或专门隔离区",
        "地面未作硬化处理",
        "附近道路运输车通行不便",
        "大件垃圾未有序码放",
        "周围存在环境卫生死角",
        "小于6平米",
        "内容不规范（灭火器材不合格）",
        "内容不规范（无灭火器）",
        "大件与装修或生活垃圾混放",
    ),
    "装修垃圾投放点设置": (
        "公示牌信息错误",
        "未设置公示牌",
        "未设置围挡或专门隔离区",
        "地面未作硬化处理",
        "附近道路运输车通行不便",
        "未袋装并有序码放",
        "周边存在环境卫生死角",
        "小于6平米",
        "内容不规范（灭火器材不合格）",
        "内容不规范（无灭火器）",
        "装修与大件或生活垃圾混放",
        "无装修垃圾备案",
    ),
}
COMMUNITY_SETTING_LEGACY_SUFFIX_ALIASES = {
    "内容不规范（无灭火器）": "内容不规范（灭火器材不合格）",
}
NOTICE_BOARD_PREFIX_MAP = {
    "小区公示牌": "小区公示牌",
    "投放点公示牌": "投放点公示牌",
}

NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
}


@dataclass
class LedgerRow:
    row_number: int
    category: str
    street: str
    place: str
    indicator1: str
    indicator2: str
    indicator3: str
    problem: str
    image_paths: list[Path] = field(default_factory=list)
    created_time: str = ""
    report_time: str = ""
    problem_image_paths_by_column: dict[int, list[Path]] = field(default_factory=dict)


@dataclass
class RowImage:
    row_number: int
    column_number: int
    media_name: str
    suffix: str
    data: bytes


@dataclass
class EmbeddedImage:
    media_name: str
    suffix: str
    data: bytes


@dataclass
class StationSection:
    title: str
    station_no: str
    problem_summary: str
    images: list[Path] = field(default_factory=list)


@dataclass
class ResidentDeliverySection:
    summary: str = ""
    error_images: list[Path] = field(default_factory=list)


@dataclass
class CommunityOverallItem:
    index: int
    title: str
    text: str = ""
    images: list[Path] = field(default_factory=list)


@dataclass
class CommunitySection:
    index_cn: str
    name: str
    overall_problem_summary: str
    overall_intro: str = ""
    promo_images: list[Path] = field(default_factory=list)
    promo_text: str = ""
    notice_board_images: list[Path] = field(default_factory=list)
    notice_board_text: str = ""
    community_litter_text: str = ""
    community_litter_images: list[Path] = field(default_factory=list)
    is_pure_box_room: bool = False
    stations: list[StationSection] = field(default_factory=list)
    resident_delivery: ResidentDeliverySection | None = None

    @property
    def overall_items(self) -> list[CommunityOverallItem]:
        items = [
            CommunityOverallItem(
                index=1,
                title="小区宣传氛围",
                text=self.promo_text if self.promo_text and self.promo_text != "无问题" else "",
                images=self.promo_images,
            ),
            CommunityOverallItem(index=2, title="小区公示牌", images=self.notice_board_images),
        ]
        if self.community_litter_text:
            items.append(
                CommunityOverallItem(
                    index=len(items) + 1,
                    title="小区垃圾乱堆乱放、投放不规范现象",
                    text=self.community_litter_text,
                    images=self.community_litter_images,
                )
            )
        if self.is_pure_box_room:
            items.append(CommunityOverallItem(index=len(items) + 1, title="装修垃圾投放点设置", text="预约收集，集中密闭运输。"))
            items.append(CommunityOverallItem(index=len(items) + 1, title="大件垃圾投放点设置", text="预约收集，集中密闭运输。"))
        return items


@dataclass
class UnitSection:
    index_cn: str
    name: str
    overall_problem_summary: str
    has_promo: bool = False
    promo_text: str = ""
    promo_images: list[Path] = field(default_factory=list)
    container_problem_summary: str = "无问题。"
    container_images: list[Path] = field(default_factory=list)


@dataclass
class StreetReport:
    street: str
    communities: list[CommunitySection] = field(default_factory=list)
    restaurants: list[UnitSection] = field(default_factory=list)
    social_units: list[UnitSection] = field(default_factory=list)


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return re.sub(r"\s+", "", text).strip()


def display_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return re.sub(r"[ \t\u3000]+", "", text).strip()


def display_time_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return re.sub(r"[ \t\u3000]+", " ", text).strip()


def load_ledger_rows(
    path: Path,
    include_images: bool = False,
    image_source_path: Path | None = None,
) -> list[LedgerRow]:
    if path.suffix.lower() == ".xlsx":
        return _load_xlsx_rows(path, include_images=include_images, image_source_path=image_source_path)
    return _load_excel_com_rows(path, include_images=include_images)


def _load_xlsx_rows(
    path: Path,
    include_images: bool = False,
    image_source_path: Path | None = None,
) -> list[LedgerRow]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [normalize_text(sheet.cell(2, col).value) for col in range(1, sheet.max_column + 1)]
    columns = _header_columns(headers)
    rows: list[LedgerRow] = []
    for row_number in range(3, sheet.max_row + 1):
        rows.append(
            LedgerRow(
                row_number=row_number,
                category=display_text(sheet.cell(row_number, columns["category"]).value),
                street=display_text(sheet.cell(row_number, columns["street"]).value),
                place=display_text(sheet.cell(row_number, columns["place"]).value),
                indicator1=display_text(sheet.cell(row_number, columns["indicator1"]).value),
                indicator2=display_text(sheet.cell(row_number, columns["indicator2"]).value),
                indicator3=display_text(sheet.cell(row_number, columns["indicator3"]).value),
                problem=display_text(sheet.cell(row_number, columns["problem"]).value),
                created_time=display_time_text(sheet.cell(row_number, columns["created_time"]).value) if columns.get("created_time") else "",
                report_time=display_time_text(sheet.cell(row_number, columns["report_time"]).value) if columns.get("report_time") else "",
            )
        )
    rows = [row for row in rows if any((row.category, row.street, row.place, row.problem))]
    if include_images:
        _attach_images(path, sheet.title, rows, image_source_path=image_source_path)
    return rows


def _load_excel_com_rows(path: Path, include_images: bool = False) -> list[LedgerRow]:
    import win32com.client

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(str(path.resolve()), ReadOnly=True)
        sheet = workbook.Worksheets(1)
        used = sheet.UsedRange
        headers = [normalize_text(sheet.Cells(2, col).Text) for col in range(1, used.Columns.Count + 1)]
        columns = _header_columns(headers)
        rows: list[LedgerRow] = []
        for row_number in range(3, used.Rows.Count + 1):
            rows.append(
                LedgerRow(
                    row_number=row_number,
                    category=display_text(sheet.Cells(row_number, columns["category"]).Text),
                    street=display_text(sheet.Cells(row_number, columns["street"]).Text),
                    place=display_text(sheet.Cells(row_number, columns["place"]).Text),
                    indicator1=display_text(sheet.Cells(row_number, columns["indicator1"]).Text),
                    indicator2=display_text(sheet.Cells(row_number, columns["indicator2"]).Text),
                    indicator3=display_text(sheet.Cells(row_number, columns["indicator3"]).Text),
                    problem=display_text(sheet.Cells(row_number, columns["problem"]).Text),
                    created_time=display_time_text(sheet.Cells(row_number, columns["created_time"]).Text) if columns.get("created_time") else "",
                    report_time=display_time_text(sheet.Cells(row_number, columns["report_time"]).Text) if columns.get("report_time") else "",
                )
            )
        workbook.Close(False)
        workbook = None
        rows = [row for row in rows if any((row.category, row.street, row.place, row.problem))]
        if include_images:
            image_root = path.parent / "extracted_images" / path.stem
            _attach_row_images(rows, extract_excel_com_row_images(path), image_root)
        return rows
    finally:
        if "workbook" in locals() and workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        try:
            excel.Quit()
        except Exception:
            pass


def _attach_images(
    path: Path,
    sheet_name: str,
    rows: list[LedgerRow],
    image_source_path: Path | None = None,
) -> None:
    image_source_path = image_source_path.resolve() if image_source_path else path
    image_root = path.parent / "extracted_images" / image_source_path.stem
    images: list[RowImage] = []
    images = extract_xlsx_row_images(path, sheet_name)
    _attach_row_images(rows, images, image_root)


def _attach_row_images(rows: list[LedgerRow], images: Iterable[RowImage], image_root: Path) -> None:
    image_root.mkdir(parents=True, exist_ok=True)
    row_numbers = [row.row_number for row in rows]
    row_map = {row.row_number: row for row in rows}
    counters: dict[int, int] = defaultdict(int)
    seen: set[str] = set()
    for image in images:
        row_number = _nearest_data_row(image.row_number, row_numbers)
        row = row_map.get(row_number) if row_number is not None else None
        if row is None:
            continue
        digest = hashlib.sha1(image.data).hexdigest()
        unique_key = f"{row.row_number}:{digest}:{image.media_name}"
        if unique_key in seen:
            continue
        seen.add(unique_key)
        counters[row.row_number] += 1
        output = image_root / f"row_{row.row_number}_{counters[row.row_number]}_{digest[:10]}{image.suffix}"
        output.write_bytes(image.data)
        row.image_paths.append(output)
        row.problem_image_paths_by_column.setdefault(image.column_number, []).append(output)


def is_outside_bucket_row(row: LedgerRow) -> bool:
    return row.category.strip() == OUTSIDE_BUCKET_POINT


def parse_report_date_value(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def normalize_sentence(text: str) -> str:
    text = str(text or "").replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    text = " ".join(text.split()).strip()
    if not text:
        return ""
    if text.endswith(("。", "！", "？", ";", "；")):
        return text
    return text + "。"


def clean_outside_bucket_text(street_name: str, problem_text: str) -> str:
    text = normalize_sentence(problem_text)
    prefix = str(street_name or "").strip()
    if prefix and text.startswith(prefix):
        text = text[len(prefix) :].lstrip("，,：:、 ")
    return normalize_sentence(text)


def make_image_rows(images: list, per_row: int = 2) -> list[list]:
    return [images[index : index + per_row] for index in range(0, len(images), per_row)]


def _outside_bucket_problem_images(row: LedgerRow) -> list[Path]:
    if not row.problem_image_paths_by_column:
        return list(row.image_paths)
    images: list[Path] = []
    for column in sorted(row.problem_image_paths_by_column):
        if column in OUTSIDE_BUCKET_PROBLEM_PHOTO_COLUMNS:
            images.extend(row.problem_image_paths_by_column[column])
    return images


def extract_outside_bucket_issues(
    rows: list[LedgerRow],
    report_date: date | datetime | None = None,
    street_filter: str | None = None,
) -> list[dict[str, object]]:
    target_date = report_date.date() if isinstance(report_date, datetime) else report_date
    target_street = (street_filter or "").strip()
    issues: list[dict[str, object]] = []

    for row in rows:
        if not is_outside_bucket_row(row):
            continue
        row_date = parse_report_date_value(row.created_time) or parse_report_date_value(row.report_time)
        if target_date and row_date and row_date != target_date:
            continue
        street_name = row.street.strip()
        if target_street and street_name != target_street:
            continue
        problem_text = row.problem.strip()
        if not problem_text or normalize_text(problem_text) in {"无问题", "无"}:
            continue
        clean_text = clean_outside_bucket_text(street_name, problem_text)
        if not clean_text:
            continue
        images = _outside_bucket_problem_images(row)
        issues.append(
            {
                "street_name": street_name,
                "street": street_name,
                "text": normalize_sentence(problem_text),
                "clean_text": clean_text,
                "images": images,
                "image_paths": images,
                "image_rows": make_image_rows(images, 2),
            }
        )
    return issues


def extract_xls_original_row_images(xls_path: Path, xlsx_path: Path, sheet_name: str) -> list[RowImage]:
    anchored_images = extract_xlsx_row_images(xlsx_path, sheet_name)
    original_images = extract_xls_embedded_images(xls_path)
    if not anchored_images and original_images:
        return extract_xls_row_images_from_biff(
            xls_path,
            image_columns=_xlsx_image_columns(xlsx_path, sheet_name),
        )
    if not anchored_images or not original_images:
        return []

    original_hashes = [(_image_hash(image.data), image) for image in original_images]
    if any(hash_value is None for hash_value, _ in original_hashes):
        return _pair_xls_images_by_order(anchored_images, original_images)

    replaced: list[RowImage] = []
    used_original_indexes: set[int] = set()
    for anchored in anchored_images:
        thumb_hash = _image_hash(anchored.data)
        if thumb_hash is None:
            replaced.append(anchored)
            continue
        best_distance = 10**9
        best_image: EmbeddedImage | None = None
        best_index: int | None = None
        for index, (original_hash, original) in enumerate(original_hashes):
            if index in used_original_indexes:
                continue
            if original_hash is None:
                continue
            distance = (thumb_hash ^ original_hash).bit_count()
            if distance < best_distance:
                best_distance = distance
                best_image = original
                best_index = index
        if best_image is None or best_index is None or best_distance > 110:
            replaced.append(anchored)
            continue
        used_original_indexes.add(best_index)
        replaced.append(
            RowImage(
                row_number=anchored.row_number,
                column_number=anchored.column_number,
                media_name=f"{anchored.media_name}|{best_image.media_name}",
                suffix=best_image.suffix,
                data=best_image.data,
            )
        )
    return replaced


def extract_xls_row_images_from_biff(path: Path, image_columns: set[int] | None = None) -> list[RowImage]:
    try:
        import olefile
    except ImportError:
        return []

    try:
        ole = olefile.OleFileIO(str(path))
    except Exception:
        return []
    try:
        workbook_data = b""
        for stream in ole.listdir(streams=True, storages=False):
            if stream[-1] in {"Workbook", "Book"}:
                workbook_data = ole.openstream(stream).read()
                break
    finally:
        ole.close()
    if not workbook_data:
        return []

    anchors = _scan_biff_image_anchors(workbook_data)
    if image_columns:
        anchors = [(row, column) for row, column in anchors if column in image_columns]
    original_images = extract_xls_embedded_images(path)
    if not anchors or not original_images:
        return []

    row_images: list[RowImage] = []
    for index, (row_number, column_number) in enumerate(anchors):
        if index >= len(original_images):
            break
        original = original_images[index]
        row_images.append(
            RowImage(
                row_number=row_number,
                column_number=column_number,
                media_name=f"xls_anchor_{index + 1}|{original.media_name}",
                suffix=original.suffix,
                data=original.data,
            )
        )
    return row_images


def _xlsx_image_columns(path: Path, sheet_name: str) -> set[int]:
    try:
        with zipfile.ZipFile(path) as archive:
            sheets = _workbook_sheet_map(archive)
            sheet_path = sheets.get(sheet_name)
            if not sheet_path:
                return FALLBACK_IMAGE_COLUMNS
            return _image_columns_for_sheet(archive, sheet_path)
    except Exception:
        return FALLBACK_IMAGE_COLUMNS


def _scan_biff_image_anchors(data: bytes) -> list[tuple[int, int]]:
    anchors: list[tuple[int, int]] = []
    for index in range(0, max(0, len(data) - 30)):
        if data[index + 2 : index + 4] != b"\x10\xf0":
            continue
        length = struct.unpack_from("<I", data, index + 4)[0]
        if length < 18 or length > 32 or index + 8 + length > len(data):
            continue
        payload = data[index + 8 : index + 8 + length]
        try:
            _flag, col1, _dx1, row1, _dy1, col2, _dx2, row2, _dy2 = struct.unpack_from("<HHHHHHHHH", payload, 0)
        except struct.error:
            continue
        if 0 <= row1 < 5000 and 0 <= col1 < 300 and row2 >= row1 and col2 >= col1:
            anchors.append((row1 + 1, col1 + 1))
    return anchors


def _pair_xls_images_by_order(anchored_images: list[RowImage], original_images: list[EmbeddedImage]) -> list[RowImage]:
    paired: list[RowImage] = []
    for index, anchored in enumerate(anchored_images):
        if index >= len(original_images):
            paired.append(anchored)
            continue
        original = original_images[index]
        paired.append(
            RowImage(
                row_number=anchored.row_number,
                column_number=anchored.column_number,
                media_name=f"{anchored.media_name}|{original.media_name}",
                suffix=original.suffix,
                data=original.data,
            )
        )
    return paired


def extract_xls_embedded_images(path: Path) -> list[EmbeddedImage]:
    try:
        import olefile
    except ImportError:
        return []

    images: list[EmbeddedImage] = []
    try:
        ole = olefile.OleFileIO(str(path))
    except Exception:
        return []
    try:
        for stream in ole.listdir(streams=True, storages=False):
            if stream[-1] not in {"Workbook", "Book"}:
                continue
            data = ole.openstream(stream).read()
            images.extend(_scan_image_blobs(data, "/".join(stream)))
    finally:
        ole.close()
    return images


def _scan_image_blobs(data: bytes, stream_name: str) -> list[EmbeddedImage]:
    images: list[EmbeddedImage] = []
    index = 0
    counters: dict[str, int] = defaultdict(int)
    signatures = ((b"\xff\xd8\xff", ".jpg"), (b"\x89PNG\r\n\x1a\n", ".png"))
    while index < len(data):
        starts = [(pos, suffix) for signature, suffix in signatures if (pos := data.find(signature, index)) != -1]
        if not starts:
            break
        start, suffix = min(starts, key=lambda item: item[0])
        if suffix == ".png":
            end = data.find(b"IEND\xaeB`\x82", start)
            if end == -1:
                index = start + 8
                continue
            end += 8
        else:
            end = data.find(b"\xff\xd9", start + 3)
            if end == -1:
                index = start + 3
                continue
            end += 2
        blob = data[start:end]
        if _valid_problem_photo_blob(blob):
            counters[suffix] += 1
            images.append(
                EmbeddedImage(
                    media_name=f"{stream_name}_image_{counters[suffix]}",
                    suffix=suffix,
                    data=blob,
                )
            )
        index = end
    return images


def _valid_problem_photo_blob(data: bytes) -> bool:
    try:
        from PIL import Image

        with Image.open(BytesIO(data)) as image:
            width, height = image.size
            image.verify()
    except Exception:
        return False
    return width >= 300 or height >= 300


def _image_hash(data: bytes) -> int | None:
    try:
        from io import BytesIO
        from PIL import Image, ImageOps

        with Image.open(BytesIO(data)) as image:
            image = ImageOps.exif_transpose(image).convert("L").resize((16, 16))
            pixels = list(image.getdata())
    except Exception:
        return None
    average = sum(pixels) / len(pixels)
    bits = 0
    for pixel in pixels:
        bits = (bits << 1) | int(pixel >= average)
    return bits


def extract_excel_com_row_images(path: Path, sheet_name: str | None = None) -> list[RowImage]:
    try:
        import win32com.client
    except ImportError:
        return []

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    images: list[RowImage] = []
    workbook = None
    try:
        workbook = excel.Workbooks.Open(str(path.resolve()), ReadOnly=True, UpdateLinks=0, AddToMru=False)
        sheet = workbook.Worksheets(sheet_name) if sheet_name else workbook.Worksheets(1)
        image_columns = _excel_image_columns(sheet)
        original_images = extract_xls_embedded_images(path)
        original_index = 0
        shape_count = int(sheet.Shapes.Count)
        for index in range(1, shape_count + 1):
            shape = sheet.Shapes.Item(index)
            try:
                row_number = int(shape.TopLeftCell.Row)
                column_number = int(shape.TopLeftCell.Column)
            except Exception:
                continue
            if column_number not in image_columns:
                continue
            if original_index >= len(original_images):
                break
            original = original_images[original_index]
            original_index += 1
            images.append(
                RowImage(
                    row_number=row_number,
                    column_number=column_number,
                    media_name=f"excel_shape_{index}|{original.media_name}",
                    suffix=original.suffix,
                    data=original.data,
                )
            )
    except Exception:
        return []
    finally:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        try:
            excel.Quit()
        except Exception:
            pass
    return images


def _excel_image_columns(sheet) -> set[int]:
    try:
        max_columns = int(sheet.UsedRange.Columns.Count)
    except Exception:
        max_columns = 200
    matches: set[int] = set()
    for column in range(1, max_columns + 1):
        try:
            value = normalize_text(sheet.Cells(1, column).Text)
        except Exception:
            continue
        if value == "问题照片":
            matches.add(column)
            try:
                merge_area = sheet.Cells(1, column).MergeArea
                start = int(merge_area.Column)
                count = int(merge_area.Columns.Count)
                matches.update(range(start, start + count))
            except Exception:
                pass
    return matches or FALLBACK_IMAGE_COLUMNS


def extract_xlsx_row_images(path: Path, sheet_name: str) -> list[RowImage]:
    images: list[RowImage] = []
    with zipfile.ZipFile(path) as archive:
        sheets = _workbook_sheet_map(archive)
        sheet_path = sheets.get(sheet_name)
        if not sheet_path:
            return images
        image_columns = _image_columns_for_sheet(archive, sheet_path)
        sheet_rels_path = f"{Path(sheet_path).parent.as_posix()}/_rels/{Path(sheet_path).name}.rels"
        sheet_rels = _read_relationships(archive, sheet_rels_path)
        sheet_root = ET.fromstring(archive.read(sheet_path))
        drawing = sheet_root.find("main:drawing", NS)
        if drawing is None:
            return images
        drawing_rel_id = drawing.attrib.get(f"{{{NS['r']}}}id")
        drawing_target = sheet_rels.get(drawing_rel_id or "")
        if not drawing_target:
            return images
        drawing_path = _rel_target_to_zip_path(Path(sheet_path).parent.as_posix(), drawing_target)
        drawing_rels_path = f"{Path(drawing_path).parent.as_posix()}/_rels/{Path(drawing_path).name}.rels"
        drawing_rels = _read_relationships(archive, drawing_rels_path)
        drawing_root = ET.fromstring(archive.read(drawing_path))

        for anchor_tag in ("xdr:twoCellAnchor", "xdr:oneCellAnchor"):
            for anchor in drawing_root.findall(anchor_tag, NS):
                from_node = anchor.find("xdr:from", NS)
                blip = anchor.find(".//a:blip", NS)
                if from_node is None or blip is None:
                    continue
                row_node = from_node.find("xdr:row", NS)
                col_node = from_node.find("xdr:col", NS)
                rel_id = blip.attrib.get(f"{{{NS['r']}}}embed")
                if row_node is None or col_node is None or not rel_id:
                    continue
                row_number = int(row_node.text or "0") + 1
                column_number = int(col_node.text or "0") + 1
                if column_number not in image_columns:
                    continue
                media_target = drawing_rels.get(rel_id)
                if not media_target:
                    continue
                media_path = _rel_target_to_zip_path(Path(drawing_path).parent.as_posix(), media_target)
                if media_path not in archive.namelist():
                    continue
                images.append(
                    RowImage(
                        row_number=row_number,
                        column_number=column_number,
                        media_name=media_path,
                        suffix=Path(media_path).suffix.lower() or ".bin",
                        data=archive.read(media_path),
                    )
                )
    return sorted(images, key=lambda image: (image.row_number, image.column_number, image.media_name))


def _image_columns_for_sheet(archive: zipfile.ZipFile, sheet_path: str) -> set[int]:
    shared_strings = _shared_strings(archive)
    sheet_root = ET.fromstring(archive.read(sheet_path))
    matches: set[int] = set()
    for cell in sheet_root.findall(".//main:sheetData/main:row[@r='1']/main:c", NS):
        value = normalize_text(_cell_text(cell, shared_strings))
        if value != "问题照片":
            continue
        column = _column_index_from_cell_ref(cell.attrib.get("r", ""))
        if column:
            matches.add(column)

    if not matches:
        return FALLBACK_IMAGE_COLUMNS

    image_columns = set(matches)
    merge_cells = sheet_root.find("main:mergeCells", NS)
    if merge_cells is not None:
        for merge_cell in merge_cells.findall("main:mergeCell", NS):
            ref = merge_cell.attrib.get("ref", "")
            bounds = _merge_bounds(ref)
            if not bounds:
                continue
            min_col, min_row, max_col, max_row = bounds
            if min_row <= 1 <= max_row and any(min_col <= column <= max_col for column in matches):
                image_columns.update(range(min_col, max_col + 1))
    return image_columns


def _shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    strings: list[str] = []
    for si in root.findall("main:si", NS):
        strings.append("".join(text.text or "" for text in si.findall(".//main:t", NS)))
    return strings


def _cell_text(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(text.text or "" for text in cell.findall(".//main:t", NS))
    value = cell.find("main:v", NS)
    if value is None or value.text is None:
        return ""
    if cell_type == "s":
        try:
            return shared_strings[int(value.text)]
        except (ValueError, IndexError):
            return ""
    return value.text


def _merge_bounds(ref: str) -> tuple[int, int, int, int] | None:
    if ":" not in ref:
        return None
    start, end = ref.split(":", 1)
    start_col, start_row = _cell_coordinate(start)
    end_col, end_row = _cell_coordinate(end)
    if not all((start_col, start_row, end_col, end_row)):
        return None
    return start_col, start_row, end_col, end_row


def _cell_coordinate(ref: str) -> tuple[int, int]:
    match = re.fullmatch(r"([A-Z]+)(\d+)", ref.upper())
    if not match:
        return 0, 0
    return _column_index_from_letters(match.group(1)), int(match.group(2))


def _column_index_from_cell_ref(ref: str) -> int:
    match = re.match(r"([A-Z]+)", ref.upper())
    return _column_index_from_letters(match.group(1)) if match else 0


def _column_index_from_letters(letters: str) -> int:
    value = 0
    for char in letters:
        value = value * 26 + ord(char) - ord("A") + 1
    return value


def _nearest_data_row(image_row: int, data_rows: list[int]) -> int | None:
    if not data_rows:
        return None
    index = bisect.bisect_right(data_rows, image_row)
    if index:
        return data_rows[index - 1]
    return data_rows[0]


def _rel_target_to_zip_path(base_dir: str, target: str) -> str:
    target_path = (Path(base_dir) / target).as_posix()
    parts: list[str] = []
    for part in target_path.split("/"):
        if part in ("", "."):
            continue
        if part == "..":
            if parts:
                parts.pop()
        else:
            parts.append(part)
    return "/".join(parts)


def _read_relationships(archive: zipfile.ZipFile, rels_path: str) -> dict[str, str]:
    if rels_path not in archive.namelist():
        return {}
    root = ET.fromstring(archive.read(rels_path))
    relationships: dict[str, str] = {}
    for rel in root.findall("rel:Relationship", NS):
        rel_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if rel_id and target:
            relationships[rel_id] = target
    return relationships


def _workbook_sheet_map(archive: zipfile.ZipFile) -> dict[str, str]:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    workbook_rels = _read_relationships(archive, "xl/_rels/workbook.xml.rels")
    sheets: dict[str, str] = {}
    for sheet in workbook_root.findall("main:sheets/main:sheet", NS):
        name = sheet.attrib.get("name")
        rel_id = sheet.attrib.get(f"{{{NS['r']}}}id")
        target = workbook_rels.get(rel_id or "")
        if name and target:
            sheets[name] = _rel_target_to_zip_path("xl", target)
    return sheets


def _header_columns(headers: list[str]) -> dict[str, int]:
    def find(name: str, occurrence: int = 1, required: bool = True) -> int | None:
        seen = 0
        for index, header in enumerate(headers, start=1):
            if header == name:
                seen += 1
                if seen == occurrence:
                    return index
        if not required:
            return None
        raise ValueError(f"Missing required header: {name}")

    return {
        "category": find("2级点位"),
        "street": find("3级点位"),
        "place": find("4级点位"),
        "indicator1": find("1级指标"),
        "indicator2": find("2级指标"),
        "indicator3": find("3级指标"),
        "problem": find("具体问题"),
        "created_time": find("创建时间", required=False),
        "report_time": find("案件上报时间", required=False),
    }


def is_no_problem(text: str) -> bool:
    normalized = normalize_text(text)
    if not normalized:
        return True
    if _is_inspection_count_text(normalized):
        return True
    no_problem_markers = (
        "无问题",
        "良好，未发现问题",
        "良好,未发现问题",
        "未发现问题",
        "有宣传氛围",
        "均有设置",
        "设施齐全",
        "已核实",
        "不是箱房小区",
        "该小区是纯箱房小区",
        "该小区是纯厢房小区",
        "没有智能可回收垃圾箱",
        "无智能可回收垃圾箱",
    )
    if normalized.isdigit():
        return True
    if normalized in {"有", "无", "五", "投放正确", "正常运行"}:
        return True
    return any(marker in normalized for marker in no_problem_markers) and not _has_problem_keyword(normalized)


def _has_problem_keyword(text: str) -> bool:
    return any(marker in text for marker in PROBLEM_KEYWORDS)


def clean_problem_text(text: str) -> str:
    text = display_text(text)
    if _has_container_count_segment(text):
        return ""
    text = re.sub(r"^无问题[，,。；;、（）()本小区今天一共检查了0-9个容器纯箱房小区]*", "", text)
    text = re.sub(r"（?本小区.*?容器。?）?", "", text)
    text = re.sub(r"本小区(?:今天)?(?:一共)?检查了?\d+个容器", "", text)
    text = re.sub(r"数量\d+个", "", text)
    text = re.sub(r"(?:本小区|小区)?(?:有|共)?\d+个(?:垃圾桶|容器)", "", text)
    text = re.sub(r"\d+个(?:垃圾桶|容器)", "", text)
    text = text.replace("。", "").replace("；", "")
    text = _strip_station_references(text)
    text = text.strip("：:，,。；;、 ")
    if is_no_problem(text):
        return ""
    return text


def _has_container_count_segment(text: str) -> bool:
    text = display_text(text)
    return bool(
        re.search(r"(?:容器数量|各容器数量|检查[^，,。；;、]*容器数量?)\d+个", text)
        or re.search(r"容器检查\d+个", text)
        or re.search(r"容器\d+个", text)
        or re.search(r"\d+个容器", text)
    )


def _is_inspection_count_text(text: str) -> bool:
    return bool(
        re.fullmatch(r"(?:本小区|小区)?(?:今天)?(?:一共)?(?:检查了?|有|共)?\d+个(?:垃圾桶|容器)", text)
    )


def effective_problem_text(row: LedgerRow) -> str:
    if _indicator3_overrides_problem_as_no_problem(row):
        return ""
    community_problem = _community_indicator_problem_text(row)
    if community_problem:
        return community_problem
    problem = clean_problem_text(row.problem)
    indicator3 = clean_problem_text(row.indicator3)
    if "小区公示牌" in row.indicator2:
        return indicator3
    if "小区宣传引导" in row.indicator2 or row.indicator2 in {"宣传引导", "宣传引导情况"}:
        return indicator3
    if "居民自主投放" in row.indicator2 and (
        "投放错误" in row.problem
        or "不准确" in row.problem
        or "投放错误" in row.indicator3
        or "不准确" in row.indicator3
    ):
        return "居民自主投放不准确"
    if "宣传引导" in row.indicator2 and ("没有看到" in row.problem or "无宣传" in row.problem):
        return "无宣传氛围"
    if is_no_problem(row.indicator3) and problem and _has_problem_keyword(problem):
        return _canonical_problem_text(problem)
    if problem == "周边不洁" and "桶站周边不洁" in indicator3:
        return "桶站周边不洁"
    if problem.isdigit() and not indicator3.isdigit():
        return indicator3
    return problem


def community_overall_problem_text(row: LedgerRow) -> str:
    if _indicator3_overrides_problem_as_no_problem(row):
        return ""
    if _is_community_litter_row(row):
        return COMMUNITY_LITTER_INDICATOR3
    if _is_community_scattered_bin_row(row):
        return COMMUNITY_SCATTERED_BIN_INDICATOR3
    if _is_resident_error(row):
        return "居民自主投放不准确"
    community_problem = _community_indicator_problem_text(row)
    if community_problem:
        return community_problem
    indicator3_raw = display_text(row.indicator3)
    problem = clean_problem_text(row.problem)
    if indicator3_raw and is_no_problem(indicator3_raw):
        if problem and _has_problem_keyword(problem):
            return _canonical_problem_text(problem)
        return ""
    indicator3 = clean_problem_text(row.indicator3)
    if "小区公示牌" in row.indicator2:
        return _prefixed_notice_board_problem(row.indicator2, indicator3)
    if indicator3:
        return indicator3
    return problem


def _canonical_problem_text(problem: str) -> str:
    if "周边" in problem and "不洁" in problem:
        return "桶站周边不洁"
    if "满冒" in problem:
        return "桶站满冒"
    if "站外摆桶" in problem or "桶外摆" in problem or "垃圾桶外摆" in problem:
        return "站外摆桶"
    if "破损" in problem or "脏污" in problem:
        return "桶站破损、脏污"
    return problem


def _community_indicator_problem_text(row: LedgerRow) -> str:
    indicator2 = display_text(row.indicator2).strip("：:，,。；;、 ")
    indicator3 = display_text(row.indicator3).strip("：:，,。；;、 ")
    problem = display_text(row.problem).strip("：:，,。；;、 ")
    if indicator2 in COMMUNITY_EXISTENCE_INDICATOR_PROBLEM_TEXT:
        if normalize_text(indicator3) == "无":
            return COMMUNITY_EXISTENCE_INDICATOR_PROBLEM_TEXT[indicator2]
        return ""
    if indicator2 in COMMUNITY_SETTING_INDICATOR_SUFFIXES:
        if is_no_problem(indicator3):
            problem_text = _expanded_community_setting_problem(indicator2, problem)
            return problem_text if problem_text else ""
        return _expanded_community_setting_problem(indicator2, indicator3) or clean_problem_text(indicator3)
    return ""


def _expanded_community_setting_problem(indicator2: str, text: str) -> str:
    text = display_text(text).strip("：:，,。；;、 ")
    if not text:
        return ""
    if is_no_problem(text):
        return ""
    if indicator2 == "装修垃圾投放点设置" and "无装修垃圾备案" in text:
        return "无装修垃圾备案"
    for suffix in COMMUNITY_SETTING_INDICATOR_SUFFIXES.get(indicator2, ()):
        expanded_suffix = COMMUNITY_SETTING_LEGACY_SUFFIX_ALIASES.get(suffix, suffix)
        expanded = f"{indicator2[:-2]}{expanded_suffix}"
        if text == suffix or text == expanded or suffix in text or expanded_suffix in text:
            return expanded
    return ""


def _prefixed_notice_board_problem(indicator2: str, text: str) -> str:
    text = clean_problem_text(text)
    if not text:
        return ""
    for indicator_key, prefix in NOTICE_BOARD_PREFIX_MAP.items():
        if indicator_key not in indicator2:
            continue
        if text.startswith(prefix):
            return text
        if text == "无公示牌":
            return f"无{prefix}"
        if text == "公示牌信息错误":
            return f"{prefix}信息错误"
        if text == "未设置公示牌":
            return f"未设置{prefix}"
        if text == "公示牌破损":
            return f"{prefix}破损"
        if text == "公示牌信息不准确":
            return f"{prefix}信息不准确"
    return text


def _indicator3_overrides_problem_as_no_problem(row: LedgerRow) -> bool:
    indicator2 = display_text(row.indicator2)
    indicator3 = normalize_text(row.indicator3)
    return (
        ("宣传引导" in indicator2 and indicator3 == "有宣传氛围")
        or ("小区公示牌" in indicator2 and indicator3 in {"良好，未发现问题", "良好,未发现问题"})
        or (_is_resident_delivery_row(row) and indicator3 == "投放正确")
    )


def problem_with_count(text: str) -> str:
    text = clean_problem_text(text)
    if not text:
        return ""
    if re.search(r"\d+处$", text):
        return text
    if text.endswith("一处"):
        return text[:-2] + "1处"
    if text.endswith("1处"):
        return text
    return text + "1处"


def problem_with_count_from_row(row: LedgerRow) -> str:
    text = effective_problem_text(row)
    if not text:
        return ""
    return problem_with_count(text)


def summarize_problem_texts(texts: Iterable[str]) -> str:
    problems: list[str] = []
    seen: set[str] = set()
    for text in texts:
        if is_no_problem(text):
            continue
        problem = problem_with_count(text)
        if problem and problem not in seen:
            seen.add(problem)
            problems.append(problem)
    if not problems:
        return "无问题。"
    if len(problems) == 1:
        return f"（1）{problems[0]}。"
    return "；".join(f"（{index}）{problem}" for index, problem in enumerate(problems, start=1)) + "。"


def summarize_problem_rows(rows: Iterable[LedgerRow]) -> str:
    ordered: list[str] = []
    counts: dict[str, int] = {}
    for row in rows:
        if is_no_problem(row.problem) and is_no_problem(row.indicator3):
            continue
        problem = effective_problem_text(row)
        if not problem:
            continue
        if problem not in counts:
            ordered.append(problem)
            counts[problem] = 0
        counts[problem] += _problem_weight(row, problem)
    problems = [_format_problem_count(problem, counts[problem]) for problem in _sort_community_problem_labels(ordered)]
    if not problems:
        return "无问题。"
    if len(problems) == 1:
        return f"（1）{problems[0]}。"
    return "；".join(f"（{index}）{problem}" for index, problem in enumerate(problems, start=1)) + "。"


def _sort_community_problem_labels(labels: list[str]) -> list[str]:
    order = {
        "桶站满冒": 0,
    }
    return sorted(labels, key=lambda label: (order.get(label, len(order)), labels.index(label)))


def _format_problem_count(problem: str, count: int) -> str:
    problem = clean_problem_text(problem)
    if not problem:
        return ""
    if re.search(r"\d+处$", problem):
        return problem
    if problem.endswith("一处"):
        problem = problem[:-2]
    elif problem.endswith("1处"):
        problem = problem[:-2]
    return f"{problem}{count}处"


def _problem_weight(row: LedgerRow, problem: str) -> int:
    if (
        problem == "居民自主投放不准确"
        and "西绒线13号院" in row.place
        and "其他投厨余" in row.problem
    ):
        return 2
    if problem == "居民自主投放不准确":
        return 1
    return _problem_count_from_problem(row.problem) or 1


def _problem_count_from_problem(text: str) -> int | None:
    text = display_text(text)
    matches = re.findall(r"([0-9一二三四五六七八九十]+)处", text)
    if not matches:
        return None
    return _parse_count_token(matches[-1])


def build_street_report(rows: list[LedgerRow], street: str) -> StreetReport:
    street_rows = [row for row in rows if row.street == street and not _is_ignored_row(row)]
    report = StreetReport(street=street)
    grouped: dict[str, list[LedgerRow]] = defaultdict(list)
    for row in street_rows:
        grouped[row.category].append(row)

    report.communities = _build_communities(grouped.get(COMMUNITY_CATEGORY, []))
    report.restaurants = _build_units(grouped.get(RESTAURANT_CATEGORY, []), RESTAURANT_CATEGORY)
    report.social_units = _build_units(grouped.get(SOCIAL_UNIT_CATEGORY, []), SOCIAL_UNIT_CATEGORY)
    return report


def _group_by_place(rows: list[LedgerRow]) -> dict[str, list[LedgerRow]]:
    grouped: dict[str, list[LedgerRow]] = {}
    for row in rows:
        grouped.setdefault(row.place, []).append(row)
    return grouped


def _is_ignored_row(row: LedgerRow) -> bool:
    return display_text(row.indicator2).strip() == "检查时段"


def _build_communities(rows: list[LedgerRow]) -> list[CommunitySection]:
    communities: list[CommunitySection] = []
    for index, (place, place_rows) in enumerate(_group_by_place(rows).items(), start=1):
        is_pure_box_room = _is_pure_box_room(place, place_rows)
        overall_rows = _overall_problem_rows(place_rows)
        section = CommunitySection(
            index_cn=chinese_numeral(index),
            name=_display_place_name(place, place_rows),
            overall_problem_summary=summarize_community_problem_rows(
                overall_rows,
                prefer_station_indicators=is_pure_box_room,
            ),
            overall_intro="",
            promo_images=_images_for_indicator(place_rows, "小区宣传"),
            promo_text=_indicator_problem_text(place_rows, "小区宣传"),
            notice_board_images=_images_for_indicator(place_rows, "小区公示牌"),
            notice_board_text=_notice_board_problem_text(place_rows),
            community_litter_text=_community_litter_text(place_rows),
            community_litter_images=_community_litter_images(place_rows),
            is_pure_box_room=is_pure_box_room,
            stations=_build_pure_box_station_sections(place_rows) if is_pure_box_room else _build_station_sections(place_rows),
            resident_delivery=_build_resident_delivery(place_rows),
        )
        _attach_unassigned_community_images(section, place_rows)
        communities.append(section)
    return communities


def _overall_problem_rows(rows: list[LedgerRow]) -> list[LedgerRow]:
    return [row for row in rows if not _is_ignored_row(row)]


def summarize_community_problem_rows(rows: Iterable[LedgerRow], prefer_station_indicators: bool = False) -> str:
    rows = [row for row in rows if not _is_ignored_row(row)]
    ordered: list[str] = []
    counts: dict[str, int] = {}
    for row in rows:
        if (
            not _is_resident_error(row)
            and not _community_indicator_pair_defines_problem(row)
            and is_no_problem(row.problem)
            and (
                is_no_problem(row.indicator3)
                or not _is_station_problem_indicator_row(row)
                or _is_explicit_no_problem_text(row.problem)
            )
        ):
            continue
        problem = community_overall_problem_text(row)
        if not problem:
            continue
        if problem == "居民自主投放不准确":
            if problem in counts:
                continue
            resident_error_rows = [item for item in rows if _is_resident_error(item)]
            ordered.append(problem)
            counts[problem] = _resident_inaccurate_count(resident_error_rows)
            continue
        if problem not in counts:
            ordered.append(problem)
            counts[problem] = 0
        counts[problem] += _problem_weight(row, problem)
    problems = [_format_problem_count(problem, counts[problem]) for problem in _sort_community_problem_labels(ordered)]
    if not problems:
        return "无问题。"
    if len(problems) == 1:
        return f"（1）{problems[0]}。"
    return "；".join(f"（{index}）{problem}" for index, problem in enumerate(problems, start=1)) + "。"


def _indicator3_can_define_problem(row: LedgerRow) -> bool:
    if "小区公示牌" in row.indicator2 and not is_no_problem(row.indicator3):
        return True
    if "宣传引导" in row.indicator2 and not is_no_problem(row.indicator3):
        return True
    if _is_community_scattered_bin_row(row):
        return True
    return False


def _is_explicit_no_problem_text(text: str) -> bool:
    normalized = normalize_text(text)
    return normalized in {
        "无",
        "无问题",
        "良好，未发现问题",
        "良好,未发现问题",
        "未发现问题",
    }


def _community_indicator_pair_defines_problem(row: LedgerRow) -> bool:
    return bool(_community_indicator_problem_text(row))


def _is_positive_indicator_result(row: LedgerRow) -> bool:
    return is_no_problem(row.indicator3) and not any(
        marker in row.problem
        for marker in ("不", "无", "未", "外摆", "满冒", "混投", "错误", "破损")
    )


def _is_pure_box_room(place: str, rows: list[LedgerRow]) -> bool:
    place_text = display_text(place)
    return (
        "纯箱房小区" in place_text
        or "纯厢房小区" in place_text
    )


COMMUNITY_LITTER_INDICATOR2 = "小区内环境"
COMMUNITY_LITTER_INDICATOR3 = "小区内有垃圾乱堆乱放、环境脏乱现象"
COMMUNITY_SCATTERED_BIN_INDICATOR3 = "散桶"


def _is_community_litter_row(row: LedgerRow) -> bool:
    return (
        display_text(row.indicator2) == COMMUNITY_LITTER_INDICATOR2
        and display_text(row.indicator3) == COMMUNITY_LITTER_INDICATOR3
        and not _community_litter_problem_says_no(row.problem)
    )


def _is_community_scattered_bin_row(row: LedgerRow) -> bool:
    return (
        display_text(row.indicator2) == COMMUNITY_LITTER_INDICATOR2
        and display_text(row.indicator3) == COMMUNITY_SCATTERED_BIN_INDICATOR3
    )


def _community_litter_problem_says_no(text: str) -> bool:
    normalized = normalize_text(text)
    if not normalized:
        return False
    return normalized in {"无", "无问题"} or any(
        marker in normalized
        for marker in ("未发现", "无乱堆乱放", "无垃圾乱堆乱放", "没有乱堆乱放", "没有垃圾乱堆乱放")
    )


def _community_litter_rows(rows: list[LedgerRow]) -> list[LedgerRow]:
    return [row for row in rows if not _is_ignored_row(row) and _is_community_litter_row(row)]


def _community_litter_text(rows: list[LedgerRow]) -> str:
    return COMMUNITY_LITTER_INDICATOR3 if _community_litter_rows(rows) else ""


def _community_litter_images(rows: list[LedgerRow]) -> list[Path]:
    return _collect_images(_community_litter_rows(rows))


def _build_station_sections(rows: list[LedgerRow]) -> list[StationSection]:
    rows = [row for row in rows if not _is_ignored_row(row)]
    station_rows = _station_rows_with_explicit_number(rows)
    if not station_rows:
        return _build_default_station_sections(rows)
    sections = _build_station_sections_from_rows(station_rows, _non_resident_rows(rows))
    unnumbered_rows = [
        row
        for row in _default_station_problem_rows(rows)
        if row not in station_rows and not station_numbers_from_problem(row.problem)
    ]
    if unnumbered_rows:
        _merge_rows_into_default_station_section(sections, unnumbered_rows, rows)
    return sections


def _is_station_setting_row(row: LedgerRow) -> bool:
    if _is_ignored_row(row):
        return False
    if _is_resident_delivery_row(row):
        return False
    return bool(station_numbers_from_problem(row.problem))


def _station_rows_with_explicit_number(rows: list[LedgerRow]) -> list[LedgerRow]:
    rows_with_station_no = [row for row in rows if _is_station_setting_row(row)]
    preferred_rows = [row for row in rows_with_station_no if _is_station_problem_indicator_row(row)]
    if preferred_rows:
        return sorted(preferred_rows, key=_station_indicator_priority)
    return rows_with_station_no


def _is_station_problem_indicator_row(row: LedgerRow) -> bool:
    indicator = display_text(row.indicator2)
    return any(keyword in indicator for keyword in _station_problem_indicator_keywords())


def _station_problem_indicator_keywords() -> tuple[str, ...]:
    return (
        "投放点环境",
        "投放点环境（容器检查）",
        "容器检查",
        "容器成组设置",
        "容器品类成组设置",
        "容器品类设置",
        "桶站便利性措施",
        "投放点公示牌",
        "早、晚高峰",
        "投放时段开盖",
        "灭蝇蚊、地面防滑设备",
        "灭蚊蝇、地面防滑设备",
        "智能回收箱是否正常运行",
        "收集车辆",
    )


def _station_indicator_priority(row: LedgerRow) -> tuple[int, int]:
    if "投放点环境" in display_text(row.indicator2) or "容器检查" in display_text(row.indicator2):
        return (0, row.row_number)
    return (1, row.row_number)


def _build_default_station_sections(rows: list[LedgerRow]) -> list[StationSection]:
    default_rows = _default_station_problem_rows(rows)
    return [
        StationSection(
            title="1号桶站设置情况",
            station_no="1",
            problem_summary=summarize_station_problem_rows(default_rows),
            images=_collect_station_images("1", _default_station_image_rows(rows, default_rows)),
        )
    ]


def _default_station_problem_rows(rows: list[LedgerRow]) -> list[LedgerRow]:
    station_rows = [
        row
        for row in rows
        if _is_station_problem_indicator_row(row)
        and not _is_resident_delivery_row(row)
        and station_problem_text_from_row(row)
    ]
    return sorted(station_rows, key=_station_indicator_priority)


def _default_station_image_rows(rows: list[LedgerRow], problem_rows: list[LedgerRow]) -> list[LedgerRow]:
    if problem_rows:
        return problem_rows
    return [row for row in rows if _is_default_station_environment_row(row)]


def _is_default_station_environment_row(row: LedgerRow) -> bool:
    indicator = display_text(row.indicator2)
    return any(
        keyword in indicator
        for keyword in (
            "投放点环境",
            "智能回收箱是否正常运行",
            "收集车辆",
        )
    )


def _build_station_sections_from_rows(station_rows: list[LedgerRow], fallback_rows: list[LedgerRow]) -> list[StationSection]:
    grouped: dict[str, list[LedgerRow]] = defaultdict(list)
    for row in station_rows:
        for station_no in station_numbers_from_problem(row.problem):
            grouped[station_no].append(row)
    sections: list[StationSection] = []
    for station_no, group in grouped.items():
        summary = summarize_station_problem_rows(group)
        images = _images_with_minimum(station_no, group, fallback_rows, minimum=3)
        sections.append(
            StationSection(
                title=f"{station_no}号桶站设置情况",
                station_no=station_no,
                problem_summary=summary,
                images=images,
            )
        )
    return _sort_station_sections(sections)


def _merge_rows_into_default_station_section(
    sections: list[StationSection],
    rows: list[LedgerRow],
    fallback_rows: list[LedgerRow],
) -> None:
    summary = summarize_station_problem_rows(rows)
    if summary == "无问题":
        return
    target = next((section for section in sections if section.station_no == "1"), None)
    if target is None:
        sections.insert(
            0,
            StationSection(
                title="1号桶站设置情况",
                station_no="1",
                problem_summary=summary,
                images=_images_with_minimum("1", rows, fallback_rows, minimum=3),
            ),
        )
        return
    if target.problem_summary == "无问题":
        target.problem_summary = summary
    else:
        target.problem_summary = "、".join(
            part for part in (target.problem_summary, summary) if part and part != "无问题"
        )
    target.images = _dedupe_path_list(target.images + _images_with_minimum("1", rows, fallback_rows, minimum=3))


def _dedupe_path_list(paths: list[Path]) -> list[Path]:
    result: list[Path] = []
    seen: set[Path] = set()
    for path in paths:
        if path in seen:
            continue
        seen.add(path)
        result.append(path)
    return result


def _sort_station_sections(sections: list[StationSection]) -> list[StationSection]:
    if any(_station_sort_key(section.station_no) is None for section in sections):
        return sections
    return sorted(sections, key=lambda section: _station_sort_key(section.station_no) or (0, ""))


def _station_sort_key(station_no: str) -> tuple[int, str] | None:
    match = re.fullmatch(r"(\d+)([A-Z]?)", station_no)
    if not match:
        return None
    return int(match.group(1)), match.group(2)


def _build_pure_box_station_sections(rows: list[LedgerRow]) -> list[StationSection]:
    rows = [row for row in rows if not _is_ignored_row(row)]
    station_rows = _station_rows_with_explicit_number(rows)
    if station_rows:
        return _build_station_sections_from_rows(station_rows, _non_resident_rows(rows))
    default_sections = _build_default_station_sections(rows)
    if default_sections:
        return default_sections
    return [StationSection(title="1号桶站设置情况", station_no="1", problem_summary="无问题", images=[])]


def _indicator_problem_text(rows: list[LedgerRow], keyword: str) -> str:
    return summarize_station_problem_rows(row for row in rows if keyword in row.indicator2)


def _notice_board_problem_text(rows: list[LedgerRow]) -> str:
    problems: list[str] = []
    seen: set[str] = set()
    for row in rows:
        if "小区公示牌" not in row.indicator2:
            continue
        problem = _prefixed_notice_board_problem(row.indicator2, row.indicator3)
        if not problem or problem in seen:
            continue
        seen.add(problem)
        problems.append(problem)
    if not problems:
        return "无问题"
    if len(problems) == 1:
        return problems[0]
    return "；".join(f"（{index}）{problem}" for index, problem in enumerate(problems, start=1)) + "。"


def summarize_station_problem_texts(texts: Iterable[str]) -> str:
    summary = summarize_problem_texts(texts)
    if summary == "无问题。":
        return "无问题"
    return re.sub(r"^（1）(.+?)。$", r"\1", summary)


def summarize_station_problem_rows(rows: Iterable[LedgerRow]) -> str:
    problems: list[str] = []
    seen: set[str] = set()
    for row in rows:
        problem = station_problem_text_from_row(row)
        if not problem or problem in seen:
            continue
        seen.add(problem)
        problems.append(problem)
    if not problems:
        return "无问题"
    if len(problems) == 1:
        return problems[0]
    return "、".join(problems)


def station_problem_text_from_row(row: LedgerRow) -> str:
    if is_no_problem(row.problem) and is_no_problem(row.indicator3):
        return ""
    indicator3_raw = display_text(row.indicator3)
    if indicator3_raw and is_no_problem(indicator3_raw):
        problem = clean_station_indicator_text(row.problem)
        if problem and _has_problem_keyword(problem) and (
            _is_station_problem_indicator_row(row) or station_numbers_from_problem(row.problem)
        ):
            return _canonical_problem_text(problem)
        return ""
    indicator3 = clean_station_indicator_text(row.indicator3)
    if indicator3:
        return _prefixed_notice_board_problem(row.indicator2, indicator3)
    problem = clean_station_indicator_text(row.problem)
    if problem:
        return _prefixed_notice_board_problem(row.indicator2, problem)
    return _prefixed_notice_board_problem(row.indicator2, clean_station_indicator_text(effective_problem_text(row)))


def clean_station_indicator_text(text: str) -> str:
    text = clean_problem_text(text)
    text = re.sub(r"\d+处$", "", text)
    text = text.removesuffix("一处")
    return text.strip("：:，,。；;、 ")


def station_number_from_problem(text: str) -> str | None:
    numbers = station_numbers_from_problem(text)
    return numbers[0] if numbers else None


def station_numbers_from_problem(text: str) -> list[str]:
    text = display_text(text)
    numbers: list[str] = []
    seen: set[str] = set()

    def add_number(raw: str) -> None:
        normalized = _normalize_station_no(re.sub(r"号$", "", raw.strip()))
        if normalized and normalized not in seen:
            seen.add(normalized)
            numbers.append(normalized)

    element = _station_number_element_pattern()
    separator = _station_number_separator_pattern()
    station_ref = rf"(?P<group>{element}(?:{separator}{element})*)\s*[）)]?\s*(?:号)?桶站"
    for match in re.finditer(station_ref, text):
        group = match.group("group")
        for raw in re.split(separator, group):
            add_number(raw)
    bare_number_ref = rf"^[（(]?\s*(?P<number>{_station_number_core_pattern()})\s*(?:号)?\s*[）)]?\s*(?=无问题|周边|满冒|外摆|混投|散桶|不洁|未|无|有害|厨余|其他|可回收|桶|垃圾|公示牌|投放|破损)"
    match = re.search(bare_number_ref, text)
    if match:
        add_number(match.group("number"))
    return numbers


def _strip_station_references(text: str) -> str:
    element = _station_number_element_pattern()
    separator = _station_number_separator_pattern()
    station_ref = rf"[（(]?\s*(?:{element}(?:{separator}{element})*)\s*[）)]?\s*(?:号)?桶站\s*[）)]?"
    matches = list(re.finditer(station_ref, text))
    if matches:
        return text[matches[-1].end():]
    bare_number_ref = rf"^[（(]?\s*{_station_number_core_pattern()}\s*(?:号)?\s*[）)]?\s*(?=无问题|周边|满冒|外摆|混投|散桶|不洁|未|无|有害|厨余|其他|可回收|桶|垃圾|公示牌|投放|破损)"
    match = re.search(bare_number_ref, text)
    if match:
        return text[match.end():]
    return text


def _station_number_element_pattern() -> str:
    return rf"(?:{_station_number_core_pattern()})(?:号)?"


def _station_number_core_pattern() -> str:
    arabic_part = r"0*\d+[A-Za-z]?"
    hyphenated_arabic = rf"{arabic_part}(?:-{arabic_part})+"
    chinese_part = r"[一二三四五六七八九十]+[A-Za-z]?"
    return rf"(?:{hyphenated_arabic}|{arabic_part}|{chinese_part})"


def _station_number_separator_pattern() -> str:
    return r"\s*(?:、|,|，|\.|．|/|和|及|与|\+)\s*"


def _normalize_station_no(station_no: str) -> str:
    if "-" in station_no:
        return "-".join(_normalize_station_no(part) for part in station_no.split("-") if part)
    match = re.fullmatch(r"0*(\d+)([A-Za-z]?)", station_no)
    if match:
        suffix = match.group(2).upper()
        return f"{int(match.group(1))}{suffix}"
    return station_no


def _build_resident_delivery(rows: list[LedgerRow]) -> ResidentDeliverySection | None:
    resident_rows = [row for row in rows if _is_resident_delivery_row(row) and not _is_ignored_row(row)]
    if not resident_rows:
        return None
    error_rows = [
        row
        for row in resident_rows
        if _is_resident_error(row)
    ]
    if not error_rows:
        return None
    inaccurate = _resident_inaccurate_count(error_rows)
    return ResidentDeliverySection(summary=f"居民投放共5个，其中不准确{inaccurate}个。", error_images=_collect_images(error_rows))


def _resident_inaccurate_count(rows: list[LedgerRow]) -> int:
    if len(rows) == 1:
        count = _resident_inaccurate_count_from_problem(rows[0].problem)
        if count is not None:
            return count
    return len(rows)


def _resident_inaccurate_count_from_problem(text: str) -> int | None:
    text = display_text(text)
    matches = re.findall(r"([0-9一二三四五六七八九十]+)[处个]", text)
    if not matches:
        return None
    return _parse_count_token(matches[-1])


def _parse_count_token(token: str) -> int:
    if token.isdigit():
        return int(token)
    numerals = {
        "一": 1,
        "二": 2,
        "三": 3,
        "四": 4,
        "五": 5,
        "六": 6,
        "七": 7,
        "八": 8,
        "九": 9,
        "十": 10,
    }
    if token == "十":
        return 10
    if token.startswith("十"):
        return 10 + numerals.get(token[1:], 0)
    if "十" in token:
        left, right = token.split("十", 1)
        return numerals.get(left, 0) * 10 + numerals.get(right, 0)
    return numerals.get(token, 1)


def _is_resident_error(row: LedgerRow) -> bool:
    if _indicator3_overrides_problem_as_no_problem(row):
        return False
    if not _is_resident_delivery_row(row):
        return False
    return (
        "投放错误" in row.problem
        or "不准确" in row.problem
        or "投放错误" in row.indicator3
        or "不准确" in row.indicator3
    )


def _is_resident_delivery_row(row: LedgerRow) -> bool:
    return "居民自主投放" in row.indicator2 or "居民投放" in row.indicator2


def _non_resident_rows(rows: Iterable[LedgerRow]) -> list[LedgerRow]:
    return [row for row in rows if not _is_resident_delivery_row(row)]


def summarize_unit_problem_rows(
    rows: Iterable[LedgerRow],
    category: str,
    numbered: bool = True,
    include_counts: bool = True,
) -> str:
    ordered: list[str] = []
    counts: dict[str, int] = {}
    for row in rows:
        problem = unit_problem_text_from_row(row, category)
        if not problem:
            continue
        if problem not in counts:
            ordered.append(problem)
            counts[problem] = 0
        counts[problem] += _unit_problem_weight(row)
    problems = [
        f"{problem}{counts[problem]}处" if include_counts else problem
        for problem in ordered
    ]
    if not problems:
        return "无问题。" if numbered else "无问题"
    if not numbered:
        return "、".join(problems)
    if len(problems) == 1:
        return f"（1）{problems[0]}。"
    return "；".join(f"（{index}）{problem}" for index, problem in enumerate(problems, start=1)) + "。"


def unit_problem_text_from_row(row: LedgerRow, category: str) -> str:
    if _unit_row_is_no_problem(row, category):
        return ""
    indicator2 = _unit_indicator_key(category, _unit_indicator_text(row.indicator2))
    indicator3 = _unit_indicator_text(row.indicator3)
    children = UNIT_INDICATOR_CHILDREN.get(category, {}).get(indicator2)

    if indicator3 and children and indicator3 in children:
        if indicator3 in UNIT_GENERIC_PROBLEM_RESULTS_USE_PARENT:
            return indicator2
        return indicator3
    child_from_problem = _unit_child_problem_from_text(category, indicator2, row.problem)
    if child_from_problem:
        return child_from_problem
    if indicator2 in UNIT_INDICATOR_CHILDREN.get(category, {}) and not children:
        return indicator2
    if indicator2 in UNIT_INDICATOR_CHILDREN.get(category, {}) and not indicator3:
        return indicator2
    if indicator3 and not _unit_indicator_result_is_no_problem(category, indicator2, indicator3):
        if indicator3 in UNIT_GENERIC_PROBLEM_RESULTS_USE_PARENT:
            return indicator2
        return indicator3
    return clean_problem_text(row.problem)


def _unit_row_is_no_problem(row: LedgerRow, category: str) -> bool:
    indicator2 = _unit_indicator_key(category, _unit_indicator_text(row.indicator2))
    indicator3 = _unit_indicator_text(row.indicator3)
    if indicator2 in UNIT_IGNORED_INDICATORS.get(category, set()):
        return True
    if indicator2 and indicator2 in UNIT_GENERIC_NO_PROBLEM_RESULTS:
        return True
    if indicator3 and _unit_indicator_result_is_no_problem(category, indicator2, indicator3):
        return True
    if _unit_problem_text_is_no_problem(row, category, indicator2, indicator3):
        return True
    if not indicator3 and normalize_text(row.problem) in UNIT_GENERIC_NO_PROBLEM_RESULTS:
        return True
    return False


def _unit_indicator_result_is_no_problem(category: str, indicator2: str, indicator3: str) -> bool:
    if indicator3 in UNIT_GENERIC_NO_PROBLEM_RESULTS:
        return True
    explicit_results = UNIT_NO_PROBLEM_RESULTS.get(category, {}).get(indicator2, set())
    if indicator3 in explicit_results:
        return True
    return _unit_indicator_pair_is_no_problem(indicator2, indicator3)


def _unit_indicator_pair_is_no_problem(indicator2: str, indicator3: str) -> bool:
    if not indicator2 or not indicator3:
        return False
    if indicator2.startswith(("无", "未")) and indicator3 in UNIT_POSITIVE_RESULT_VALUES:
        return True
    if indicator2.startswith(("无", "未")) and indicator3.startswith(("有", "已")):
        return True
    if any(keyword in indicator2 for keyword in UNIT_BAD_EVENT_INDICATOR_KEYWORDS) and indicator3 in UNIT_NEGATIVE_RESULT_VALUES:
        return True
    if indicator2 == "分类投放情况" and indicator3 == "投放正确":
        return True
    if "宣传引导" in indicator2 and indicator3 in {"有", "有宣传氛围"}:
        return True
    if "容器" in indicator2 and indicator3 in {"成组配置", "良好，未发现问题", "良好,未发现问题"}:
        return True
    return False


def _unit_problem_text_is_no_problem(row: LedgerRow, category: str, indicator2: str, indicator3: str) -> bool:
    problem = _unit_indicator_text(row.problem)
    if not problem:
        return False
    markers = UNIT_NO_PROBLEM_TEXT_MARKERS.get(category, {}).get(indicator2, set())
    if not markers:
        return False
    if indicator3:
        return False
    for marker in markers:
        if marker == problem:
            return True
        if len(marker) > 1 and marker in problem:
            return True
    return False


def _unit_indicator_text(value: object) -> str:
    return display_text(value).strip("：:，,。；;、 ")


def _unit_indicator_key(category: str, indicator2: str) -> str:
    return UNIT_INDICATOR_ALIASES.get(category, {}).get(indicator2, indicator2)


def _unit_child_problem_from_text(category: str, indicator2: str, text: str) -> str:
    text = display_text(text)
    for child in UNIT_INDICATOR_CHILDREN.get(category, {}).get(indicator2, ()):
        if _unit_indicator_result_is_no_problem(category, indicator2, child):
            continue
        if child in UNIT_GENERIC_PROBLEM_RESULTS_USE_PARENT:
            continue
        if child and child in text:
            return child
    return ""


def _unit_problem_weight(row: LedgerRow) -> int:
    return _unit_problem_count_from_text(row.problem) or 1


def _unit_problem_count_from_text(text: str) -> int | None:
    text = display_text(text)
    matches = re.findall(r"([0-9一二三四五六七八九十]+)[处个组]", text)
    if not matches:
        return None
    return _parse_count_token(matches[-1])


def _build_units(rows: list[LedgerRow], category: str) -> list[UnitSection]:
    units: list[UnitSection] = []
    for index, (place, place_rows) in enumerate(_group_by_place(rows).items(), start=1):
        closed_text = _closed_unit_problem_text(place_rows)
        if closed_text:
            units.append(
                UnitSection(
                    index_cn=chinese_numeral(index),
                    name=place,
                    overall_problem_summary=closed_text,
                    has_promo=True,
                    promo_text=closed_text,
                    promo_images=_collect_images(place_rows),
                    container_problem_summary=closed_text,
                    container_images=[],
                )
            )
            continue
        promo_rows = [row for row in place_rows if "宣传引导" in row.indicator2]
        container_rows = [row for row in place_rows if "容器" in row.indicator2 or "容器" in row.indicator3]
        if not container_rows:
            container_rows = [row for row in place_rows if row.image_paths]
        promo_images = _collect_images(promo_rows) if promo_rows else _stable_sample_images(place, _collect_images(place_rows), 2)
        container_images = _collect_images(container_rows)
        container_images.extend(_remaining_images(place_rows, promo_images + container_images))
        units.append(
            UnitSection(
                index_cn=chinese_numeral(index),
                name=place,
                overall_problem_summary=summarize_unit_problem_rows(place_rows, category),
                has_promo=True,
                promo_text=summarize_unit_problem_rows(promo_rows, category, numbered=False) if promo_rows else "无问题",
                container_problem_summary=summarize_unit_problem_rows(
                    container_rows,
                    category,
                    numbered=False,
                    include_counts=False,
                ),
                promo_images=_unique_images(promo_images),
                container_images=_unique_images(container_images),
            )
        )
    return units


def _closed_unit_problem_text(rows: Iterable[LedgerRow]) -> str:
    for row in rows:
        if _is_closed_unit_problem(row.problem):
            return "已关门"
    return ""


def _is_closed_unit_problem(text: str) -> bool:
    normalized = display_text(text)
    return bool(re.search(r"(已|己)?(?:关门|停业|歇业|闭店|停止营业|暂停营业)", normalized))


def _stable_sample_images(place: str, images: list[Path], count: int) -> list[Path]:
    if count <= 0 or not images:
        return []
    if len(images) <= count:
        return images
    seed = hashlib.sha1(place.encode("utf-8")).hexdigest()
    return random.Random(seed).sample(images, count)


def _collect_images(rows: Iterable[LedgerRow]) -> list[Path]:
    images: list[Path] = []
    for row in rows:
        images.extend(row.image_paths)
    return _unique_images(images)


def _images_with_minimum(
    station_no: str,
    primary_rows: Iterable[LedgerRow],
    fallback_rows: Iterable[LedgerRow],
    minimum: int,
) -> list[Path]:
    primary_images = _collect_station_images(station_no, primary_rows)
    return primary_images


def _collect_station_images(station_no: str, rows: Iterable[LedgerRow]) -> list[Path]:
    seen_first_images: set[str] = set()
    images: list[Path] = []
    for row in rows:
        row_images = list(row.image_paths)
        if not row_images:
            continue
        first_key = _image_identity(row_images[0])
        if first_key in seen_first_images:
            row_images = row_images[1:]
        else:
            seen_first_images.add(first_key)
        images.extend(row_images)
    return _unique_images(images)


def _image_identity(image: Path) -> str:
    try:
        return f"sha1:{hashlib.sha1(image.read_bytes()).hexdigest()}"
    except OSError:
        return f"path:{Path(image)}"


def _attach_unassigned_community_images(section: CommunitySection, rows: list[LedgerRow]) -> None:
    assigned = section.promo_images + section.notice_board_images + section.community_litter_images
    for station in section.stations:
        assigned.extend(station.images)
    if section.resident_delivery:
        assigned.extend(section.resident_delivery.error_images)

    remaining = _remaining_images(_non_resident_rows(rows), assigned)
    if not remaining:
        return

    if section.stations:
        return

    section.stations.append(
        StationSection(
            title="1号桶站设置情况",
            station_no="1",
            problem_summary="无问题",
            images=remaining,
        )
    )


def _remaining_images(rows: Iterable[LedgerRow], assigned: Iterable[Path]) -> list[Path]:
    assigned_keys = {_image_identity(image) for image in assigned}
    remaining: list[Path] = []
    for image in _collect_images(rows):
        if _image_identity(image) not in assigned_keys:
            remaining.append(image)
    return remaining


def _unique_images(images: Iterable[Path]) -> list[Path]:
    return list(images)


def _display_place_name(place: str, rows: list[LedgerRow]) -> str:
    if any(
        "不是箱房小区" in display_text(row.problem)
        or "不是厢房小区" in display_text(row.problem)
        or "不是箱房小区" in display_text(row.indicator3)
        or "不是厢房小区" in display_text(row.indicator3)
        for row in rows
    ):
        return f"{place}（经检查员核实，该小区不属于纯箱房小区）"
    return place





def _images_for_indicator(rows: list[LedgerRow], indicator: str) -> list[Path]:
    images: list[Path] = []
    for row in rows:
        if indicator in row.indicator2:
            images.extend(row.image_paths)
    return images


def chinese_numeral(number: int) -> str:
    values = "零一二三四五六七八九十"
    if 0 <= number <= 10:
        return values[number]
    if number < 20:
        return "十" + values[number % 10]
    return str(number)
